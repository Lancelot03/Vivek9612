import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Generator
from datetime import datetime, date
import io
import base64
import asyncio
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import uuid

class ExcelExportService:
    """Advanced Excel export service with styling, batching, and progress tracking"""
    
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
        self.export_tasks = {}  # Store export progress
    
    def create_styled_workbook(self) -> Workbook:
        """Create a new workbook with default styling"""
        wb = Workbook()
        
        # Define common styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        return wb
    
    def style_worksheet(self, ws, data_rows: int, start_row: int = 1):
        """Apply styling to worksheet"""
        
        # Style header row
        for cell in ws[start_row]:
            if cell.value:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
        
        # Style data rows
        for row in range(start_row + 1, start_row + data_rows + 1):
            for cell in ws[row]:
                cell.border = self.border
                if isinstance(cell.value, str) and len(str(cell.value)) > 50:
                    cell.alignment = self.wrap_alignment
                else:
                    cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def export_responses_advanced(self, batch_size: int = 500) -> Dict[str, Any]:
        """Advanced response export with multiple sheets and analytics"""
        
        export_id = str(uuid.uuid4())
        self.export_tasks[export_id] = {
            "status": "starting",
            "progress": 0,
            "total_steps": 6,
            "current_step": "Fetching responses",
            "start_time": datetime.utcnow()
        }
        
        try:
            # Step 1: Fetch all responses
            self.export_tasks[export_id]["current_step"] = "Fetching responses data"
            responses = await self.db.responses.find().to_list(10000)
            self.export_tasks[export_id]["progress"] = 1
            
            if not responses:
                self.export_tasks[export_id]["status"] = "error"
                self.export_tasks[export_id]["error"] = "No responses found"
                return {"error": "No responses to export", "export_id": export_id}
            
            # Step 2: Fetch invitee details for enrichment
            self.export_tasks[export_id]["current_step"] = "Enriching with invitee data"
            invitees = await self.db.invitees.find().to_list(10000)
            invitee_lookup = {inv["employeeId"]: inv for inv in invitees}
            self.export_tasks[export_id]["progress"] = 2
            
            # Step 3: Process and enrich response data
            self.export_tasks[export_id]["current_step"] = "Processing response data"
            enriched_responses = []
            
            for response in responses:
                invitee = invitee_lookup.get(response["employeeId"], {})
                
                enriched_response = {
                    "Employee ID": response.get("employeeId", ""),
                    "Employee Name": invitee.get("employeeName", "Unknown"),
                    "Cadre": invitee.get("cadre", "Not Specified"),
                    "Project Name": invitee.get("projectName", "Not Specified"),
                    "Mobile Number": response.get("mobileNumber", ""),
                    "Requires Accommodation": "Yes" if response.get("requiresAccommodation") else "No",
                    "Arrival Date": response.get("arrivalDate", ""),
                    "Departure Date": response.get("departureDate", ""),
                    "Food Preference": response.get("foodPreference", ""),
                    "Departure Time Preference": response.get("departureTimePreference", ""),
                    "Arrival Time Preference": response.get("arrivalTimePreference", ""),
                    "Special Flight Requirements": response.get("specialFlightRequirements", ""),
                    "Submission Date": response.get("submissionTimestamp").strftime('%Y-%m-%d') if response.get("submissionTimestamp") else "",
                    "Submission Time": response.get("submissionTimestamp").strftime('%H:%M:%S') if response.get("submissionTimestamp") else ""
                }
                enriched_responses.append(enriched_response)
            
            self.export_tasks[export_id]["progress"] = 3
            
            # Step 4: Create workbook with multiple sheets
            self.export_tasks[export_id]["current_step"] = "Creating Excel workbook"
            wb = self.create_styled_workbook()
            
            # Main responses sheet
            ws_responses = wb.active
            ws_responses.title = "All Responses"
            
            # Convert to DataFrame for easy manipulation
            df_responses = pd.DataFrame(enriched_responses)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df_responses, index=False, header=True):
                ws_responses.append(r)
            
            # Style the responses sheet
            self.style_worksheet(ws_responses, len(df_responses))
            self.export_tasks[export_id]["progress"] = 4
            
            # Step 5: Create analytics sheets
            self.export_tasks[export_id]["current_step"] = "Creating analytics sheets"
            
            # Accommodation analysis sheet
            ws_accommodation = wb.create_sheet("Accommodation Analysis")
            accommodation_data = []
            
            # Accommodation summary
            total_responses = len(enriched_responses)
            accommodation_yes = len([r for r in enriched_responses if r["Requires Accommodation"] == "Yes"])
            accommodation_no = total_responses - accommodation_yes
            
            accommodation_data = [
                ["Accommodation Summary", "", ""],
                ["Total Responses", total_responses, ""],
                ["Requires Accommodation", accommodation_yes, f"{(accommodation_yes/total_responses*100):.1f}%"],
                ["No Accommodation", accommodation_no, f"{(accommodation_no/total_responses*100):.1f}%"],
                ["", "", ""],
                ["Accommodation Details", "", ""],
                ["Employee ID", "Employee Name", "Arrival Date", "Departure Date"]
            ]
            
            # Add accommodation details
            for response in enriched_responses:
                if response["Requires Accommodation"] == "Yes":
                    accommodation_data.append([
                        response["Employee ID"],
                        response["Employee Name"], 
                        response["Arrival Date"],
                        response["Departure Date"]
                    ])
            
            for row in accommodation_data:
                ws_accommodation.append(row)
            
            self.style_worksheet(ws_accommodation, len(accommodation_data))
            
            # Food preferences analysis sheet
            ws_food = wb.create_sheet("Food Preferences")
            food_counts = pd.Series([r["Food Preference"] for r in enriched_responses]).value_counts()
            
            food_data = [
                ["Food Preference Analysis", "", "", ""],
                ["Food Type", "Count", "Percentage", ""],
            ]
            
            for food_type, count in food_counts.items():
                percentage = (count / total_responses * 100)
                food_data.append([food_type, count, f"{percentage:.1f}%", ""])
            
            food_data.append(["", "", "", ""])
            food_data.append(["Total Responses", total_responses, "100%", ""])
            
            for row in food_data:
                ws_food.append(row)
                
            self.style_worksheet(ws_food, len(food_data))
            
            # Project-wise breakdown sheet
            ws_projects = wb.create_sheet("Project Breakdown")
            project_counts = pd.Series([r["Project Name"] for r in enriched_responses]).value_counts()
            
            project_data = [
                ["Project-wise Response Analysis", "", "", ""],
                ["Project Name", "Responses", "Percentage", ""],
            ]
            
            for project, count in project_counts.items():
                percentage = (count / total_responses * 100)
                project_data.append([project, count, f"{percentage:.1f}%", ""])
            
            for row in project_data:
                ws_projects.append(row)
                
            self.style_worksheet(ws_projects, len(project_data))
            
            self.export_tasks[export_id]["progress"] = 5
            
            # Step 6: Generate final Excel file
            self.export_tasks[export_id]["current_step"] = "Generating final file"
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            self.export_tasks[export_id]["progress"] = 6
            self.export_tasks[export_id]["status"] = "completed"
            self.export_tasks[export_id]["end_time"] = datetime.utcnow()
            
            filename = f"PM_Connect_Comprehensive_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return {
                "export_id": export_id,
                "excel_data": excel_base64,
                "filename": filename,
                "summary": {
                    "total_responses": int(total_responses),
                    "accommodation_requests": int(accommodation_yes),
                    "food_preferences": {str(k): int(v) for k, v in food_counts.items()},
                    "project_breakdown": {str(k): int(v) for k, v in project_counts.items()},
                    "sheets_created": ["All Responses", "Accommodation Analysis", "Food Preferences", "Project Breakdown"]
                }
            }
            
        except Exception as e:
            self.export_tasks[export_id]["status"] = "error"
            self.export_tasks[export_id]["error"] = str(e)
            self.export_tasks[export_id]["end_time"] = datetime.utcnow()
            raise e
    
    async def export_invitees_with_status(self) -> Dict[str, Any]:
        """Export invitees with their response status"""
        
        export_id = str(uuid.uuid4())
        self.export_tasks[export_id] = {
            "status": "starting",
            "progress": 0,
            "total_steps": 4,
            "current_step": "Fetching invitee data",
            "start_time": datetime.utcnow()
        }
        
        try:
            # Fetch invitees and responses
            invitees = await self.db.invitees.find().to_list(10000)
            responses = await self.db.responses.find().to_list(10000)
            self.export_tasks[export_id]["progress"] = 1
            
            # Create response lookup
            response_lookup = {r["employeeId"]: r for r in responses}
            self.export_tasks[export_id]["progress"] = 2
            
            # Process invitee data
            self.export_tasks[export_id]["current_step"] = "Processing invitee data"
            invitee_data = []
            
            for invitee in invitees:
                emp_id = invitee["employeeId"]
                response = response_lookup.get(emp_id)
                
                invitee_row = {
                    "Employee ID": emp_id,
                    "Employee Name": invitee.get("employeeName", ""),
                    "Cadre": invitee.get("cadre", ""),
                    "Project Name": invitee.get("projectName", ""),
                    "Response Status": "Responded" if response else "Pending",
                    "Mobile Number": response.get("mobileNumber", "") if response else "",
                    "Accommodation": "Yes" if response and response.get("requiresAccommodation") else "No" if response else "",
                    "Food Preference": response.get("foodPreference", "") if response else "",
                    "Response Date": response.get("submissionTimestamp").strftime('%Y-%m-%d') if response and response.get("submissionTimestamp") else ""
                }
                invitee_data.append(invitee_row)
            
            self.export_tasks[export_id]["progress"] = 3
            
            # Create Excel file
            self.export_tasks[export_id]["current_step"] = "Creating Excel file"
            df = pd.DataFrame(invitee_data)
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Invitees with Status', index=False)
                
                # Style the worksheet
                workbook = writer.book
                worksheet = workbook['Invitees with Status']
                
                # Apply basic styling
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
            
            excel_bytes = excel_buffer.getvalue()
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            self.export_tasks[export_id]["progress"] = 4
            self.export_tasks[export_id]["status"] = "completed"
            self.export_tasks[export_id]["end_time"] = datetime.utcnow()
            
            filename = f"PM_Connect_Invitees_Status_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return {
                "export_id": export_id,
                "excel_data": excel_base64,
                "filename": filename,
                "summary": {
                    "total_invitees": len(invitees),
                    "responded": len([i for i in invitee_data if i["Response Status"] == "Responded"]),
                    "pending": len([i for i in invitee_data if i["Response Status"] == "Pending"])
                }
            }
            
        except Exception as e:
            self.export_tasks[export_id]["status"] = "error"
            self.export_tasks[export_id]["error"] = str(e)
            self.export_tasks[export_id]["end_time"] = datetime.utcnow()
            raise e
    
    async def export_cab_allocations(self) -> Dict[str, Any]:
        """Export cab allocations with member details"""
        
        export_id = str(uuid.uuid4())
        
        try:
            # Fetch data
            cab_allocations = await self.db.cab_allocations.find().to_list(1000)
            invitees = await self.db.invitees.find().to_list(10000)
            
            # Create invitee lookup
            invitee_lookup = {inv["employeeId"]: inv for inv in invitees}
            
            # Process cab data
            cab_data = []
            for cab in cab_allocations:
                for emp_id in cab["assignedMembers"]:
                    invitee = invitee_lookup.get(emp_id, {})
                    cab_data.append({
                        "Cab Number": cab["cabNumber"],
                        "Employee ID": emp_id,
                        "Employee Name": invitee.get("employeeName", "Unknown"),
                        "Cadre": invitee.get("cadre", ""),
                        "Project Name": invitee.get("projectName", ""),
                        "Pickup Location": cab["pickupLocation"],
                        "Pickup Time": cab["pickupTime"]
                    })
            
            # Create Excel file
            df = pd.DataFrame(cab_data)
            excel_buffer = io.BytesIO()
            df.to_excel(excel_buffer, index=False, sheet_name='Cab Allocations')
            
            excel_bytes = excel_buffer.getvalue()
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            filename = f"PM_Connect_Cab_Allocations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return {
                "export_id": export_id,
                "excel_data": excel_base64,
                "filename": filename,
                "summary": {
                    "total_cabs": len(cab_allocations),
                    "total_members": len(cab_data)
                }
            }
            
        except Exception as e:
            raise e
    
    def get_export_progress(self, export_id: str) -> Dict[str, Any]:
        """Get progress of an export task"""
        return self.export_tasks.get(export_id, {"status": "not_found"})
    
    def cleanup_completed_exports(self, max_age_hours: int = 24):
        """Clean up completed export tasks older than specified hours"""
        cutoff_time = datetime.utcnow() - pd.Timedelta(hours=max_age_hours)
        
        completed_exports = [
            export_id for export_id, task in self.export_tasks.items()
            if task.get("status") in ["completed", "error"] and 
               task.get("end_time", datetime.utcnow()) < cutoff_time
        ]
        
        for export_id in completed_exports:
            del self.export_tasks[export_id]
        
        return len(completed_exports)